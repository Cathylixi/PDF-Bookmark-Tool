"""Read the TOC Excel file and produce FILENAME -> FINAL_TITLE mapping."""

import os
import tempfile
import zipfile
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from utils.logger import get_logger
from utils.models import ExcelMappingResult, ExcelRow

logger = get_logger(__name__)

REQUIRED_COLUMNS = ["FILENAME", "FINAL_TITLE"]

MINIMAL_CORE_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"'
    ' xmlns:dc="http://purl.org/dc/elements/1.1/"'
    ' xmlns:dcterms="http://purl.org/dc/terms/"'
    ' xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
    "</cp:coreProperties>"
)


def _rebuild_xlsx_with_clean_core(src_path: str) -> str:
    """Copy ``src_path`` to a temp .xlsx whose docProps/core.xml is replaced
    by a minimal valid document so openpyxl can load workbooks that were
    written with slightly malformed ISO dates in their metadata.
    """
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="toc_sanitized_")
    os.close(fd)
    try:
        with zipfile.ZipFile(src_path, "r") as src_zip, zipfile.ZipFile(
            tmp_path, "w", zipfile.ZIP_DEFLATED
        ) as dst_zip:
            for item in src_zip.infolist():
                if item.filename == "docProps/core.xml":
                    dst_zip.writestr(item, MINIMAL_CORE_XML)
                else:
                    dst_zip.writestr(item, src_zip.read(item.filename))
        return tmp_path
    except Exception:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        raise


def _safe_load_workbook(excel_path: str):
    """Load a workbook, falling back to a sanitized copy if the original has
    unreadable document properties (e.g. invalid ISO date in docProps)."""
    try:
        return load_workbook(excel_path, data_only=True, read_only=True), None
    except ValueError as exc:
        msg = str(exc)
        if "could not read properties" in msg or "ISO datetime" in msg:
            logger.warning(
                "Workbook has malformed core properties, rebuilding: %s", msg
            )
            tmp_path = _rebuild_xlsx_with_clean_core(excel_path)
            wb = load_workbook(tmp_path, data_only=True, read_only=True)
            return wb, tmp_path
        raise


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def _locate_header_row(ws) -> Tuple[Optional[int], dict]:
    """Scan the first 10 rows to find the header row that contains FILENAME and FINAL_TITLE."""
    max_scan = min(10, ws.max_row or 0)
    for row_idx in range(1, max_scan + 1):
        header_map = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            name = _normalize_header(cell.value)
            if name in REQUIRED_COLUMNS and name not in header_map:
                header_map[name] = col_idx
        if all(col in header_map for col in REQUIRED_COLUMNS):
            return row_idx, header_map
    return None, {}


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_excel_mapping(excel_path: str) -> ExcelMappingResult:
    """Read FILENAME and FINAL_TITLE columns from the first worksheet.

    The reader tolerates extra columns and locates the header row dynamically
    within the first 10 rows.
    """
    result = ExcelMappingResult()
    logger.info("Reading Excel: %s", excel_path)

    wb, sanitized_tmp = _safe_load_workbook(excel_path)
    try:
        ws = wb[wb.sheetnames[0]]
        header_row, header_map = _locate_header_row(ws)

        if header_row is None:
            result.missing_columns = [
                col for col in REQUIRED_COLUMNS if col not in header_map
            ] or REQUIRED_COLUMNS[:]
            logger.warning("Missing required columns: %s", result.missing_columns)
            return result

        filename_col = header_map["FILENAME"]
        title_col = header_map["FINAL_TITLE"]

        seen_filenames: dict = {}
        duplicates: List[str] = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if row is None:
                continue

            filename = _cell_text(row[filename_col - 1]) if len(row) >= filename_col else ""
            title = _cell_text(row[title_col - 1]) if len(row) >= title_col else ""

            if not filename and not title:
                continue

            result.total_rows += 1

            if not filename:
                result.empty_title_rows.append(
                    ExcelRow(row_number=row_idx, filename="", title=title)
                )
                continue

            if not title:
                result.empty_title_rows.append(
                    ExcelRow(row_number=row_idx, filename=filename, title="")
                )
                continue

            if filename in seen_filenames:
                if filename not in duplicates:
                    duplicates.append(filename)
                continue

            seen_filenames[filename] = title
            result.mapping[filename] = title

        result.duplicate_filenames = duplicates
        logger.info(
            "Excel parsed: %d usable rows, %d duplicates, %d empty-title rows",
            len(result.mapping),
            len(duplicates),
            len(result.empty_title_rows),
        )
        return result
    finally:
        wb.close()
        if sanitized_tmp and os.path.exists(sanitized_tmp):
            try:
                os.remove(sanitized_tmp)
            except OSError:
                pass
